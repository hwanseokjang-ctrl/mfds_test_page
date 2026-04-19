"""배치 분석 페이지 — 엑셀/CSV 업로드 → 일괄 분석 → 결과 확인·다운로드."""
from __future__ import annotations

import io
import re as _re
from datetime import datetime
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment

from common import llm_client
from common.components import render_header, render_llm_panel, render_prompt_editor
from common.llm_process import ProcessResult
from common.page_loader import load_page_module
from page_modules.product_name_analysis import prompts as pna_prompts

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
MODULE_NAME = "product_name_analysis"

_SEG_KEYS = [f"seg_{i}" for i in range(1, 7)]
_SEG_LABELS: dict[str, str] = {
    "seg_1": "원재료명",
    "seg_2": "통칭명(C)",
    "seg_3": "식품유형명·요리명(D)",
    "seg_4": "추출물·농축액",
    "seg_5": "성분명",
    "seg_6": "맛·향(E)",
}

# ---------------------------------------------------------------------------
# 모듈 로드
# ---------------------------------------------------------------------------
config, _run, _ = load_page_module(MODULE_NAME)

# ---------------------------------------------------------------------------
# 공통 헤더 / LLM 패널 / 프롬프트 편집기
# ---------------------------------------------------------------------------
render_header("제품명 분석 — 배치", "엑셀/CSV를 업로드해 여러 제품을 한 번에 분석합니다.")
render_llm_panel()
render_prompt_editor(pna_prompts.PROMPT_CATALOG, key_prefix="batch_" + config.slug)

st.divider()


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _strip_for_excel(text: str) -> str:
    """HTML 태그·마크다운 문법 제거 — Excel 출력용 plain text 변환."""
    text = _re.sub(r'<[^>]+>', '', text)
    text = _re.sub(r'\*{1,3}([^*\n]+)\*{1,3}', r'\1', text)
    text = _re.sub(r'^#{1,6}\s+', '', text, flags=_re.MULTILINE)
    text = _re.sub(r'^>\s+', '', text, flags=_re.MULTILINE)
    text = _re.sub(r'^---\s*$', '', text, flags=_re.MULTILINE)
    text = _re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _snapshot_settings() -> dict[str, str]:
    """분석 시점의 LLM 설정을 캡처해 dict 로 반환."""
    provider_name = st.session_state.get(llm_client.SESSION_PROVIDER, "-")
    model = st.session_state.get(llm_client.SESSION_MODEL, "-")
    temperature = st.session_state.get(llm_client.SESSION_TEMPERATURE, "-")
    reasoning_effort = st.session_state.get(llm_client.SESSION_REASONING_EFFORT, "-")
    try:
        provider_info = llm_client.get_provider(provider_name)
        provider_display = provider_info.display
    except Exception:
        provider_display = provider_name
    return {
        "제공자": provider_display,
        "모델": model,
        "Temperature": str(temperature),
        "Reasoning Effort": str(reasoning_effort),
    }


def _get_seg_result(result: ProcessResult, seg_key: str) -> dict[str, Any]:
    r = result.intermediates.get(seg_key, {}).get("result", {})
    return r if isinstance(r, dict) else {}


def _overall_judgment(result: ProcessResult) -> str:
    for seg_key in _SEG_KEYS:
        if _get_seg_result(result, seg_key).get("result") == "해당":
            return "해당"
    return "해당없음"


@st.cache_data(show_spinner=False)
def _make_template_bytes() -> bytes:
    df = pd.DataFrame(columns=["제품명", "원재료명"])
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def _parse_file(data: bytes, name: str) -> pd.DataFrame:
    if name.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(data))
    return pd.read_excel(io.BytesIO(data), engine="openpyxl")


def _best_col(candidates: list[str], cols: list[str]) -> str:
    for keyword in candidates:
        for col in cols:
            if keyword in col.lower():
                return col
    return cols[0]


def _build_row_data(product_name: str, ingredients: str, extra: dict[str, str]) -> dict[str, Any]:
    """빈 row_data 템플릿 생성."""
    row_data: dict[str, Any] = {
        "제품명": product_name,
        "원재료명": ingredients,
        **extra,
        "상태": "",
        "오류": "",
        "최종판정": "",
        "분석결과": "",
        "전처리결과": "",
    }
    for seg_key in _SEG_KEYS:
        lbl = _SEG_LABELS[seg_key]
        for field in ["결과", "근거", "keyword_list", "err_keyword_list", "passed_list"]:
            row_data[f"[{lbl}] {field}"] = ""
    row_data["_result_obj"] = None
    return row_data


def _build_excel_bytes(results: list[dict[str, Any]], settings: dict[str, str] | None = None) -> bytes:
    skip = {"_result_obj"}
    excel_rows = []
    for r in results:
        row = {k: v for k, v in r.items() if k not in skip}
        if "분석결과" in row and isinstance(row["분석결과"], str):
            row["분석결과"] = _strip_for_excel(row["분석결과"])
        excel_rows.append(row)
    dl_df = pd.DataFrame(excel_rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        dl_df.to_excel(writer, sheet_name="분석결과", index=False)
        ws = writer.sheets["분석결과"]

        # 분석결과 열: 줄바꿈 텍스트 표시 + 열 너비 확보
        if "분석결과" in dl_df.columns:
            col_idx = list(dl_df.columns).index("분석결과") + 1  # 1-based
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = 80
            for row_num in range(2, len(dl_df) + 2):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 전처리결과 열: wrap_text 적용
        if "전처리결과" in dl_df.columns:
            col_idx = list(dl_df.columns).index("전처리결과") + 1
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = 60
            for row_num in range(2, len(dl_df) + 2):
                ws.cell(row=row_num, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

        # 근거 열들도 wrap_text 적용
        for seg_key in _SEG_KEYS:
            lbl = _SEG_LABELS[seg_key]
            col_name = f"[{lbl}] 근거"
            if col_name in dl_df.columns:
                col_idx = list(dl_df.columns).index(col_name) + 1
                col_letter = ws.cell(1, col_idx).column_letter
                ws.column_dimensions[col_letter].width = 40
                for row_num in range(2, len(dl_df) + 2):
                    ws.cell(row=row_num, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )

        if settings:
            meta_rows = [{"항목": k, "값": v} for k, v in settings.items()]
            meta_rows.append({"항목": "분석일시", "값": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            meta_rows.append({"항목": "분석건수", "값": str(len(results))})
            pd.DataFrame(meta_rows).to_excel(writer, sheet_name="분석설정", index=False)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 배치 실행 중 처리 (한 건씩 rerun 방식 — 중단 버튼 활성화)
# ---------------------------------------------------------------------------
_is_running = st.session_state.get("batch_running", False)

if _is_running:
    _idx = st.session_state.get("batch_idx", 0)
    _total = st.session_state.get("batch_total", 1)
    _pname = st.session_state.get("batch_current_name", "…")

    # 진행 바 + 중단 버튼
    pb_col, stop_col = st.columns([6, 1])
    with pb_col:
        progress_pct = _idx / max(_total, 1)
        st.progress(progress_pct, text=f"처리 중 **{_idx} / {_total}** — {_pname}")
    with stop_col:
        if st.button("⏹ 중단", type="secondary", width="stretch"):
            st.session_state["batch_stop_requested"] = True

    _records = st.session_state["batch_df_records"]
    _col_product = st.session_state["batch_col_product"]
    _col_ingr = st.session_state["batch_col_ingr"]
    _extra_cols = st.session_state["batch_extra_cols"]
    _partial: list[dict[str, Any]] = st.session_state["batch_partial_results"]

    _stop_requested = st.session_state.get("batch_stop_requested", False)

    if _stop_requested or _idx >= _total:
        # 완료 또는 중단
        st.session_state["batch_running"] = False
        st.session_state["batch_results"] = _partial
        st.session_state["batch_partial_results"] = []
        st.session_state["batch_was_stopped"] = _stop_requested
        st.rerun()
    else:
        # 한 건 처리
        _row = _records[_idx]
        _product_name = str(_row.get(_col_product) or "").strip()
        _ingredients = str(_row.get(_col_ingr) or "").strip()
        _extra = {c: str(_row.get(c, "")) for c in _extra_cols}

        st.session_state["batch_current_name"] = _product_name or "(빈 값)"
        _row_data = _build_row_data(_product_name, _ingredients, _extra)

        try:
            _result: ProcessResult = _run(
                {"product_name": _product_name, "ingredients": _ingredients}
            )
            _row_data["상태"] = "성공"
            _row_data["최종판정"] = _overall_judgment(_result)
            _row_data["분석결과"] = _result.final_text
            _row_data["전처리결과"] = _result.intermediates.get("preprocess", {}).get("result", "")
            _row_data["_result_obj"] = _result

            for _seg_key in _SEG_KEYS:
                _lbl = _SEG_LABELS[_seg_key]
                _seg_r = _get_seg_result(_result, _seg_key)
                _row_data[f"[{_lbl}] 결과"] = _seg_r.get("result", "")
                _row_data[f"[{_lbl}] 근거"] = _seg_r.get("reason", "")
                _row_data[f"[{_lbl}] keyword_list"] = _seg_r.get("keyword_list", "")
                _row_data[f"[{_lbl}] err_keyword_list"] = _seg_r.get("err_keyword_list", "")
                _row_data[f"[{_lbl}] passed_list"] = _seg_r.get("passed_list", "")
        except Exception as _exc:
            _row_data["상태"] = "실패"
            _row_data["오류"] = f"{type(_exc).__name__}: {_exc}"

        _partial.append(_row_data)
        st.session_state["batch_partial_results"] = _partial
        st.session_state["batch_idx"] = _idx + 1
        st.rerun()

    st.stop()  # 아래 정상 흐름 실행 차단


# ---------------------------------------------------------------------------
# 중단 완료 알림
# ---------------------------------------------------------------------------
if st.session_state.pop("batch_was_stopped", False):
    n = len(st.session_state.get("batch_results") or [])
    st.warning(f"⏹ 분석이 중단되었습니다. {n}건의 결과가 아래에 저장되어 있습니다.")


# ---------------------------------------------------------------------------
# 템플릿 다운로드
# ---------------------------------------------------------------------------
st.download_button(
    "📥 입력 템플릿 다운로드",
    _make_template_bytes(),
    file_name="분석_템플릿.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ---------------------------------------------------------------------------
# 파일 업로드
# ---------------------------------------------------------------------------
uploaded = st.file_uploader(
    "Excel / CSV 업로드",
    type=["xlsx", "xls", "csv"],
    label_visibility="collapsed",
)

if uploaded is None:
    # 결과가 이미 있으면 파일 없어도 결과 섹션까지 진행
    if not st.session_state.get("batch_results"):
        st.info("📂 분석할 파일을 업로드하거나 위 템플릿을 다운로드해 양식을 확인하세요.")
        st.stop()
else:
    # ---------------------------------------------------------------------------
    # 파싱
    # ---------------------------------------------------------------------------
    try:
        df_raw = _parse_file(uploaded.read(), uploaded.name)
    except Exception as exc:
        st.error(f"파일 파싱 실패: {exc}")
        st.stop()

    if df_raw.empty or len(df_raw.columns) < 2:
        st.error("파일에 데이터가 없거나 열이 부족합니다.")
        st.stop()

    cols = list(df_raw.columns)

    # ---------------------------------------------------------------------------
    # 컬럼 매핑
    # ---------------------------------------------------------------------------
    st.subheader("컬럼 매핑")
    c1, c2 = st.columns(2)
    with c1:
        default_pname = _best_col(["제품명", "product_name", "productname", "품명"], cols)
        col_product = st.selectbox(
            "제품명 열",
            cols,
            index=cols.index(default_pname) if default_pname in cols else 0,
        )
    with c2:
        default_ingr = _best_col(["원재료", "ingredient", "원료", "재료"], cols)
        col_ingr = st.selectbox(
            "원재료명 열",
            cols,
            index=cols.index(default_ingr) if default_ingr in cols else 0,
        )

    extra_cols = [c for c in cols if c not in (col_product, col_ingr)]

    # ---------------------------------------------------------------------------
    # 미리보기
    # ---------------------------------------------------------------------------
    st.subheader(f"미리보기 — 총 {len(df_raw):,}건")
    preview_cols = [col_product, col_ingr] + extra_cols[:3]
    st.dataframe(df_raw[preview_cols].head(5), width="stretch")

    if len(df_raw) > 200:
        st.warning(f"총 {len(df_raw)}건으로 분석 시간이 길어질 수 있습니다. LLM 비용도 함께 확인하세요.")

    # ---------------------------------------------------------------------------
    # 실행 버튼
    # ---------------------------------------------------------------------------
    if st.button("▶ 분석 실행", type="primary", width="stretch"):
        st.session_state.update(
            {
                "batch_running": True,
                "batch_stop_requested": False,
                "batch_was_stopped": False,
                "batch_idx": 0,
                "batch_total": len(df_raw),
                "batch_partial_results": [],
                "batch_results": None,
                "batch_settings": _snapshot_settings(),
                "batch_df_records": df_raw.to_dict("records"),
                "batch_col_product": col_product,
                "batch_col_ingr": col_ingr,
                "batch_extra_cols": extra_cols,
                "batch_current_name": "",
            }
        )
        st.rerun()


# ---------------------------------------------------------------------------
# 결과 렌더 (파일 유무와 무관하게 표시)
# ---------------------------------------------------------------------------
results: list[dict[str, Any]] | None = st.session_state.get("batch_results")
batch_settings: dict[str, str] | None = st.session_state.get("batch_settings")

if not results:
    st.stop()

success_rows = [r for r in results if r["상태"] == "성공"]
failed_rows = [r for r in results if r["상태"] == "실패"]
applicable_rows = [r for r in success_rows if r["최종판정"] == "해당"]

st.divider()

# 분석 설정 요약
if batch_settings:
    with st.expander("⚙️ 분석 시점 설정", expanded=False):
        cols_s = st.columns(len(batch_settings))
        for col, (k, v) in zip(cols_s, batch_settings.items()):
            col.metric(k, v)

# KPI 카드
st.subheader("📊 분석 요약")
k1, k2, k3, k4 = st.columns(4)
k1.metric("총 분석", f"{len(results):,}건")
k2.metric("성공", f"{len(success_rows):,}건")
k3.metric(
    "함량표시 해당",
    f"{len(applicable_rows):,}건",
    f"{len(applicable_rows)/max(len(success_rows),1)*100:.1f}%",
)
k4.metric("실패(오류)", f"{len(failed_rows):,}건")

# 세분화별 해당 비율 바차트
st.subheader("세분화별 해당 건수")
seg_counts = {
    _SEG_LABELS[k]: sum(
        1 for r in success_rows if r.get(f"[{_SEG_LABELS[k]}] 결과") == "해당"
    )
    for k in _SEG_KEYS
}
bar_df = pd.DataFrame(
    {"세분화": list(seg_counts.keys()), "해당 건수": list(seg_counts.values())}
).set_index("세분화")
st.bar_chart(bar_df)

st.divider()

# 결과 다운로드
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
st.download_button(
    "📥 결과 다운로드 (Excel)",
    _build_excel_bytes(results, batch_settings),
    file_name=f"배치분석결과_{timestamp}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)

# 결과 요약 테이블
st.subheader("결과 테이블")
table_display_cols = ["제품명", "최종판정", "상태"] + [
    f"[{_SEG_LABELS[k]}] 결과" for k in _SEG_KEYS
]
display_df = pd.DataFrame(
    [{c: r.get(c, "") for c in table_display_cols} for r in results]
)


def _color_judgment(val: str) -> str:
    if val == "해당":
        return "background-color: #fff3cd; color: #856404;"
    if val == "해당없음":
        return "background-color: #d1e7dd; color: #0f5132;"
    if val == "실패":
        return "background-color: #f8d7da; color: #842029;"
    return ""


styled = display_df.style.map(
    _color_judgment,
    subset=["최종판정"] + [f"[{_SEG_LABELS[k]}] 결과" for k in _SEG_KEYS],
)
st.dataframe(styled, width="stretch", height=400)

# 행별 상세 보기
st.subheader("행별 상세 결과")
for i, row in enumerate(results):
    icon = "✅" if row["상태"] == "성공" else "❌"
    _jp = row["최종판정"]
    if _jp == "해당":
        judgment_badge = "⚠️ 해당"
    elif _jp == "해당없음":
        judgment_badge = "✅ 해당없음"
    else:
        judgment_badge = _jp
    with st.expander(f"{icon} {i+1}. {row['제품명']} — {judgment_badge}"):
        if row["상태"] == "실패":
            st.error(f"오류: {row['오류']}")
            continue

        st.markdown(f"**분석 결론**\n\n{row['분석결과']}", unsafe_allow_html=True)
        st.divider()

        result_obj: ProcessResult | None = row.get("_result_obj")
        if result_obj is None:
            continue

        st.markdown("**세분화별 결과**")
        for seg_key in _SEG_KEYS:
            lbl = _SEG_LABELS[seg_key]
            seg_r = _get_seg_result(result_obj, seg_key)
            seg_result = seg_r.get("result", "-")
            icon_seg = "🔴" if seg_result == "해당" else "⚪"
            with st.expander(f"{icon_seg} {lbl}: {seg_result}"):
                if seg_r.get("reason"):
                    st.caption(f"**근거:** {seg_r['reason']}")
                if seg_r.get("keyword_list"):
                    st.caption(f"🔑 keyword_list: `{seg_r['keyword_list']}`")
                if seg_r.get("err_keyword_list"):
                    st.caption(f"⚠️ err_keyword_list: `{seg_r['err_keyword_list']}`")
                if seg_r.get("passed_list"):
                    st.caption(f"✅ passed_list: `{seg_r['passed_list']}`")
